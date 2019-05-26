
import openpyxl
from datetime import datetime, timedelta
import pymssql
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill, numbers
import os

conn_43 = pymssql.connect(host = "10.28.7.43",user = "bond",password = 'bond',database = "qbdb")
conn_43_creditdb = pymssql.connect(host = "10.28.7.43",user = "bond",password = 'bond',database = "creditdb")

#  表头字段：调整时间	发行人名称 主体资信级别上调	主体评级展望上调	评级机构	调整原因	行业	调整方向
creditratingagency={
    "1": "标准普尔评级服务公司",
    "2": "中诚信国际信用评级有限责任公司",
    "3": "上海远东资信评估有限公司",
    "4": "上海新世纪资信评估投资服务有限公司",
    "5": "联合资信评估有限公司",
    "6": "大公国际资信评估有限公司",
    "7": "联合信用评级有限公司",
    "8": "辽宁省资信评估有限公司",
    "9": "长城资信评估有限公司",
    "10": "云南省资信评估事务所",
    "11": "北京穆迪投资者服务有限公司",
    "12": "福建省资信评级委员会",
    "13": "中诚信证券评估有限公司",
    "14": "鹏元资信评估有限公司",
    "15": "中国证券监督管理委员会",
    "16": "穆迪公司",
    "17": "惠誉国际信用评级有限公司",
    "18": "中债资信评估有限责任公司",
    "19": "东方金诚国际信用评估有限公司",
    "20": "联合信用管理有限公司",
    "21": "天津海泰信用服务有限公司",
    "22": "中国证券业协会",
    "23": "中华信用评等公司",
    "24": "南京中贝国际信用管理咨询有限公司",
    "25": "北京资信评级有限公司",
    "26": "中国诚信信用管理股份有限公司",
    "27": "上海资信有限公司",
    "28": "穆迪投资者服务香港有限公司",
    "29": "标准普尔香港有限公司",
    "30": "江苏安博尔信用评估有限公司"
}

creditratingoutlook = {
    "518001000": "正面",
    "518002000": "负面",
    "518003000": "稳定",
    "518006000": "观望",
    "518007000": "待决"
}



def bondInfo(bondCode):
    cur = conn_43.cursor()
    sql="""
        SELECT  
            *
        FROM
            OPENQUERY (
            WINDNEW,
            'select a.s_info_windcode,a.b_issue_firstissue,a.s_info_name,a.b_info_couponrate,a.b_info_issuer,a.b_info_issuercode from CBondDescription a  where a.s_info_windcode = ''%s'''
        )
    """ % bondCode
    cur.execute(sql)
    _bondinfo = cur.fetchone()
    cur.close()
    return _bondinfo

def bondWithSameIssuer(companyName):
    cur = conn_43.cursor()
    sql = """
        SELECT
        *
        FROM
            OPENQUERY (
            WINDNEW,
            'select s_info_windcode from CBondIssuer b where b.S_INFO_COMPNAME = ''%s'''
        )
    """ % companyName
    cur.execute(sql)
    _bonds = cur.fetchall()
    cur.close()
    return _bonds

# 查询主体评级
def issuerCreditRating(companyName):
    cur = conn_43.cursor()
    sql="""
        SELECT  
        *
        FROM
            OPENQUERY (
            WINDNEW,
            'select ANN_DT,b_info_creditrating,b_rate_ratingoutlook,b_info_creditratingagency from CBondIssuerRating where s_info_compname=''%s'' order by ANN_DT desc'
        )
    """ % companyName
    cur.execute(sql)
    _issuercreditrating = cur.fetchall()
    cur.close()
    return _issuercreditrating

# 行业&城投 信用评级调整
def creditChange(industry, Date, changeDirection=1):
    cur = conn_43_creditdb.cursor()
    sql = """
    SELECT updatetime,issuer,creditChange,creditOutlookChange,creditAgency,comment,industry,changeDirection from CreditChangebyIndustry WHERE industry = '%s' AND changeDirection=%d AND updatetime>='%s'
    """ %(industry, changeDirection, (Date-timedelta(365)).strftime("%Y%m%d"))
    cur.execute(sql)
    return cur.fetchall()

# 行业&城投 同行业同评级券种查询
def bondInSameInd(ind, creditRating):
    cur = conn_43_creditdb.cursor()
    # sql = """
    # SELECT * from BondInSameInd where industry='%s' AND creditRating='%s' ORDER BY updatetime DESC
    # """ %(ind, creditRating)
    sql = """
        SELECT * from BondInSameInd where industry='%s' ORDER BY updatetime DESC
        """ % (ind)
    cur.execute(sql)
    _temp = cur.fetchall()
    cur.close()
    return _temp


def bondinput(bondCode, codename, industry, specialterm="", interval=1):
    today = datetime.now()
    border = Border(left=Side(border_style='thin',
                              color='FF000000'),
                    right=Side(border_style='thin',
                               color='FF000000'),
                    top=Side(border_style='thin',
                             color='FF000000'),
                    bottom=Side(border_style='thin',
                                color='FF000000')
                    )
    font = Font(name='宋体',size = 11,bold = True,italic = False,vertAlign = None,underline = 'none',strike = False, color = 'FF000000')
    alignment = Alignment(horizontal='center', vertical = 'center', text_rotation = 0, wrap_text = False, shrink_to_fit = False,indent=0)
    fill = PatternFill("solid", fgColor="FCD5B4")
    number_format = 'General'

    wb = openpyxl.load_workbook('20181010-模板 - 副本.xlsx')
    ws = wb['公式页']
    ws['A2'] = bondCode
    for cellList in ws['A4:G11']:
        for cell in cellList:
            cell.border = border

    debtinfo = bondInfo(bondCode)
    if debtinfo is None:
        return 0
    issuer = debtinfo[4]
    # codename = debtinfo[2]
    bonds = bondWithSameIssuer(issuer)
    bondsList = []
    for bond in bonds:
        bondinfo = bondInfo(bond[0])
        if bondinfo is not None and bondinfo[0][0] not in ('q', 'd'):
            bondsList.append(bondinfo)
    # df = pd.DataFrame(bondsList,columns=['s_info_windcode','b_issue_firstissue','s_info_name','b_info_couponrate','b_info_issuercode'])


    ws = wb["发行人概况"]
    # for r in dataframe_to_rows(df,index=False,header=False):
    #     ws.append(r)
    if bondsList.__len__() >1:
        bondsList.sort(key=lambda x: x[1] if x[1] is not None else '-1', reverse=True)
    for i in range(bondsList.__len__()):
        ws['A%d' % (i + 13)] = bondsList[i][0]
        ws['B%d' % (i + 13)] = bondsList[i][1]
        ws['C%d' % (i + 13)] = bondsList[i][2]
        ws['D%d' % (i + 13)] = bondsList[i][3]
        ws['E%d' % (i + 13)] = """=b_anal_ptmyear(A%d,"")"""% (i + 13)
        ws['F%d' % (i + 13)] = """=b_rate_latestcredit(A%d)"""% (i + 13)
        ws['G%d' % (i + 13)] = """=b_info_issueamount(A%d)/100000000"""% (i + 13)

    issuercreditrating = issuerCreditRating(issuer)

    # 主体评级表头
    startrow = bondsList.__len__()+18
    ws.merge_cells('A%d:D%d'%(startrow,startrow))
    ws['A%d' % startrow] = "历史主体评级"
    ws['A%d' % (startrow+1)] = "发布日期"
    ws['B%d' % (startrow + 1)] = "主体资信级别"
    ws['C%d' % (startrow + 1)] = "评级展望"
    ws['D%d' % (startrow + 1)] = "评级机构"
    newIssuerCreditRating = issuercreditrating[0][1] if issuercreditrating.__len__() > 0 else None
    for i in range(issuercreditrating.__len__()):
        ws['A%d' % (i + startrow + 2)] = issuercreditrating[i][0]
        ws['B%d' % (i + startrow + 2)] = issuercreditrating[i][1]
        ws['C%d' % (i + startrow + 2)] = creditratingoutlook[str(issuercreditrating[i][2])] if issuercreditrating[i][2] is not None else "--"
        ws['D%d' % (i + startrow + 2)] = creditratingagency[str(issuercreditrating[i][3])] if issuercreditrating[i][3] is not None else "--"
    for cellList in ws['A%d:D%d' % (startrow,startrow+1)]:
        for cell in cellList:
            cell.border = border
            cell.font = font
            cell.alignment = alignment

    for cell in ws['A11:G11'][0]:
        cell.border = border
    for cellList in ws['A2:G9']:
        for cell in cellList:
            cell.border = border

    ws = wb['发行人财务信息'] # 制作发行人财务信息表格格式
    for cellList in ws['A1:J26']:
        for cell in cellList:
            cell.border = border

    ws = wb['行业内评级变动']
    # 行业变动上调
    upList = creditChange(industry, today, 1)

    for i in range(upList.__len__()):
        ws['A%d' % (i + 3)] = upList[i][0]
        ws['A%d' % (i + 3)].number_format = 'yyyy/mm/dd'
        ws['B%d' % (i + 3)] = upList[i][1]
        ws['C%d' % (i + 3)] = upList[i][2]
        ws['C%d' % (i + 3)].fill = fill
        ws['D%d' % (i + 3)] = upList[i][3]
        ws['D%d' % (i + 3)].fill = fill
        ws['E%d' % (i + 3)] = upList[i][4]
        ws['F%d' % (i + 3)] = upList[i][5]
        ws['F%d' % (i + 3)].fill = fill

    #行业评级下降表头
    startrow=upList.__len__()+18
    ws.merge_cells('A%d:F%d'%(startrow,startrow))
    ws['A%d' % startrow] = "近一年来同行业发债企业主体评级下调情况"
    ws['A%d' % (startrow + 1)] = "调整时间"
    ws['B%d' % (startrow + 1)] = "发行人名称"
    ws['C%d' % (startrow + 1)] = "主体资信级别下调"
    ws['D%d' % (startrow + 1)] = "主体评级展望下调"
    ws['E%d' % (startrow + 1)] = "评级机构"
    ws['F%d' % (startrow + 1)] = "调整原因"

    for cellList in ws['A%d:F%d' % (startrow,startrow+1)]:
        for cell in cellList:
            cell.border = border
            cell.font = font
            cell.alignment = alignment

    # 行业变动下降
    downList = creditChange(industry, today, 0)
    for i in range(downList.__len__()):
        ws['A%d' % (i + startrow + 2)] = downList[i][0]
        ws['A%d' % (i + startrow + 2)].number_format = 'yyyy/mm/dd'
        ws['B%d' % (i + startrow + 2)] = downList[i][1]
        ws['C%d' % (i + startrow + 2)] = downList[i][2]
        ws['C%d' % (i + startrow + 2)].fill = fill
        ws['D%d' % (i + startrow + 2)] = downList[i][3]
        ws['D%d' % (i + startrow + 2)].fill = fill
        ws['E%d' % (i + startrow + 2)] = downList[i][4]
        ws['F%d' % (i + startrow + 2)] = downList[i][5]
        ws['F%d' % (i + startrow + 2)].fill = fill
    for cellList in ws['A%d:F%d' % (startrow+2, downList.__len__()+startrow+1)]:
        for cell in cellList:
            cell.border = border

    ws = wb['公式页']
    bondinsameind = bondInSameInd(industry.replace('城投-',''), newIssuerCreditRating)
    j=0
    for i in range(bondinsameind.__len__()):
        if issuer != bondinsameind[i][3]:
            creditratingList = issuerCreditRating(bondinsameind[i][3])
            if creditratingList.__len__() > 0 and creditratingList[0][1] == newIssuerCreditRating:
                ws['T%d' % (j+15)] = bondinsameind[i][1]
                ws['%s16' % chr(j+76)] = newIssuerCreditRating
                j += 1
                if j >= 7:
                    break


    ws = wb['同行业财务比较']
    for i in range(6,17,1):
        ws['C%d' % i] = """=sum(D%d:J%d)/%d"""%(i, i, j) if j > 0 else "--"

    # 文件夹命名
    if specialterm == "" or specialterm is None: filename = (today + timedelta(interval)).strftime("%Y%m%d") + "-" + codename + "-" + industry
    else: filename = (today + timedelta(interval)).strftime("%Y%m%d") + "-" + codename + "-" + industry + "-" + specialterm
    wb.save('./%s新券信评/%s.xlsx' % ((today + timedelta(interval)).strftime("%Y%m%d"), filename))
    wb.close()
    print(filename+" finished")




    return 1


if __name__ == '__main__':
    inputrows=[ #['d19012903.IB','19光大集团SCP001','综合','0.25'],
                #['d19012904.IB','19陕投集团SCP001','煤炭',''],
        ['d19031806.IB', '19渤海金控SCP001', '租赁', ''],
        ['d19031805.IB', '19巨石SCP001', '化工', ''],
        ['d19031502.IB', '19云冶SCP002', '有色', ''],
    ]
    #inputrows = input("输入要生成的新券数据:")
    interval = int(input("输入发行日距今间隔(整数):"))
    today = datetime.now()
    os.mkdir("%s新券信评" % (today + timedelta(interval)).strftime("%Y%m%d"))
    with open("CreditEstimate.txt",'r') as f:
        line = f.readline().replace('\n', '')
        while line.__len__() > 2:
            data = line.split(',')
            if data.__len__() == 4:
                bondinput(data[0], data[1], data[2], specialterm=data[3], interval=interval)
            elif data.__len__() == 3:
                bondinput(data[0], data[1], data[2], specialterm="", interval=interval)
            line = f.readline().replace('\n', '')
    print("enter any keys to exit")
    input()
    print("bye")

    # for data in inputrows:
    #     bondinput(data[0], data[1], data[2], specialterm=data[3])






